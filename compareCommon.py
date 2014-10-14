def checkParams(params, count):
	if len(params) != count:
		return False;
	return True

def checkDictKeys(dicts):
	keys = []
	for d in dicts:
		keys.append(d.keys())
	
	for k in keys:
		if not (keys[0] == k):
			return False

	return True

def isInt(val):
	try:
		int(val)
		return True
	except ValueError:
		return False

def createKeysDict(keys):
	allKeysOK = True
	for k in keys:
		if not '-' in k:
			allKeysOK = False
			break
		if not isInt(k[k.rfind('-'):]):
			allKeysOK = False
			break

	if not allKeysOK:
		return dict(zip(keys,keys))

	newKeys = [s[:s.rfind('-')+1] + '{:05}'.format(int(s[s.rfind('-')+1:])) for s in keys]
	return dict(zip(newKeys, keys))

def compare(englishDict, originalDict, correctedDict, notesDict):
	result = []
	keysDict = createKeysDict(originalDict.keys())
	for k in sorted(keysDict.keys()):
		key = keysDict[k]
		if (correctedDict[key] != originalDict[key]) or (key in notesDict):
			note = ''
			if (key in notesDict):
				note = notesDict[key];
			result.append([key, englishDict[key], originalDict[key], correctedDict[key], note])
	return result

def csvOutput(result, fileName):
	f = open(fileName, 'w')
	for row in result:
		for column in row:
			f.write(column.replace('\n',' ') + '|')
		f.write('\n')

def htmlOutput(result, fileName):
	f = open(fileName, 'w')
	f.write("<html><head><meta http-equiv=\"Content-Type\" content=\"text/html;charset=utf-8\"></head>")
	f.write("<body><table border = \"1\"><tr><th>Key</th><th>Source</th><th>Translated</th><th>Corrected</th><th>Note</th></tr>")
	for row in result:
		f.write("<tr>")
		for column in row:
			f.write("<td>"+column + '</td>')
		f.write("</tr>")
	f.write("</table></body></html>")

def excelOutput(result, fileName):
	try:
		from openpyxl import Workbook
		from openpyxl.worksheet import ColumnDimension 
		from openpyxl.worksheet import RowDimension
	except Exception:
		sys.exit("python modyle openpyxl is not installed")

	wb = Workbook()
	ws = wb.get_active_sheet()

	cellKey = ws.cell(row = 0, column = 0)
	cellKey.value = "Key"
	cellKey.style.font.bold = True
	cellOriginal = ws.cell(row = 0, column = 1)
	cellOriginal.value = "Source"
	cellOriginal.style.font.bold = True
	cellTranslated = ws.cell(row = 0, column = 2)
	cellTranslated.value = "Translated"
	cellTranslated.style.font.bold = True
	cellCorrected = ws.cell(row = 0, column = 3)
	cellCorrected.value = "Corrected"
	cellCorrected.style.font.bold = True
	cellNote = ws.cell(row = 0, column = 4)
	cellNote.value = "Note"
	cellNote.style.font.bold = True
	ws.column_dimensions["A"].width = 25
	ws.column_dimensions["B"].width = 50
	ws.column_dimensions["C"].width = 50
	ws.column_dimensions["D"].width = 50
	ws.column_dimensions["E"].width = 35

	rowIndex = 1
	for rowContent in result:
		colIndex = 0
		for cellContent in rowContent:
			cell = ws.cell(row = rowIndex, column = colIndex)
			cell.value = cellContent
			cell.style.alignment.wrap_text = True
			cell.style.height = 20
			colIndex += 1
		rowIndex += 1
		ws.row_dimensions[rowIndex].height = 50

	wb.save(fileName)